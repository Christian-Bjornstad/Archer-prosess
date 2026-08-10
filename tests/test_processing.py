from pathlib import Path

import openpyxl

from archer_processor.core import (
    FilterEngine,
    VariantProcessor,
    default_artifact_rules,
    production_rules,
)
from archer_processor.core.models import DatabaseEvidence
from archer_processor.io import ArcherTsvReader
from archer_processor.reports import ExcelReportWriter
from archer_processor.services import DatabaseSearchService, load_database_skip_keys


FIXTURE = Path(__file__).parent / "fixtures" / "sample_variants.tsv"


def test_reader_parses_archer_tsv_columns():
    variants = ArcherTsvReader().read(FIXTURE)

    assert len(variants) == 5
    assert variants[0].sample == "26OUM00001_VPM_S1_R1_001"
    assert variants[0].patient_id == "26OUM00001"
    assert variants[0].transcript == "NM_004119.2"
    assert variants[3].af == 0.5333


def test_production_rules_and_boundaries():
    variants = ArcherTsvReader().read(FIXTURE)
    FilterEngine().apply(variants)

    by_sample = {variant.patient_id: variant for variant in variants}
    assert by_sample["26OUM00001"].decision == "excluded"
    assert by_sample["26OUM00002"].decision == "excluded"
    assert by_sample["26OUM00003"].decision == "excluded"
    assert by_sample["26OUM00004"].warnings
    assert by_sample["26OUM00005"].decision == "excluded"
    assert by_sample["26OUM00005"].warnings


def test_default_artifact_catalog_comes_from_fragmentation_v2_hgvsc_column():
    rules = default_artifact_rules()

    assert len(rules) == 36
    assert len({entry["hgvsc"] for entry in rules}) == 36
    assert {entry["gene"] for entry in rules} >= {
        "ASXL1",
        "ATRX",
        "CBL",
        "CEBPA",
        "DDX41",
        "EZH2",
        "FLT3",
        "JAK2",
        "NOTCH1",
        "NPM1",
        "PTEN",
        "RUNX1",
        "SRSF2",
        "STAG2",
    }


def test_asxl1_1934dup_is_artifact_through_5_5_percent_only():
    variants = ArcherTsvReader().read(FIXTURE)
    at_threshold = variants[1]
    above_threshold = variants[2]
    at_threshold.af = 0.055
    above_threshold.af = 0.055001

    FilterEngine().apply([at_threshold, above_threshold])

    assert at_threshold.decision == "excluded"
    assert "artifact" in at_threshold.matched_rules[0]
    assert above_threshold.decision == "included"


def test_custom_artifact_rules_replace_default_artifact_list():
    variants = ArcherTsvReader().read(FIXTURE)
    rules = production_rules(
        [
            {
                "gene": "TP53",
                "hgvsc": "NM_000546.6:c.524G>A",
                "reason": "Temporary local artifact for testing.",
            }
        ]
    )
    FilterEngine(rules).apply(variants)

    by_sample = {variant.patient_id: variant for variant in variants}
    assert by_sample["26OUM00001"].decision == "included"
    assert by_sample["26OUM00004"].decision == "excluded"
    assert "artifact" in by_sample["26OUM00004"].matched_rules[0]


def test_processor_writes_excel(tmp_path):
    output = tmp_path / "review.xlsx"
    result = VariantProcessor().process(FIXTURE, "2026-07-26", output)
    ExcelReportWriter().write(result, output)

    assert output.exists()
    assert result.total_count == 5
    assert len(result.included) == 1
    assert len(result.excluded) == 4


def test_excel_export_preserves_raw_columns_and_adds_database_columns(tmp_path):
    output = tmp_path / "review.xlsx"
    result = VariantProcessor().process(FIXTURE, "2026-07-26", output)
    variant = result.variants[3]
    evidence = {
        DatabaseSearchService().variant_key(variant): [
            DatabaseEvidence("ClinVar", "found", "ClinVar summary"),
            DatabaseEvidence("gnomAD", "found", "gnomAD summary"),
            DatabaseEvidence("COSMIC", "found", "COSMIC summary"),
            DatabaseEvidence("CIViC", "found", "CIViC summary"),
            DatabaseEvidence("CancerMine", "found", "CancerMine summary"),
            DatabaseEvidence("DGIdb", "found", "DGIdb summary"),
            DatabaseEvidence("ClinGen Allele Registry", "found", "ClinGen summary"),
            DatabaseEvidence("cBioPortal", "found", "cBioPortal summary"),
            DatabaseEvidence("OncoKB", "token_required", "OncoKB token required"),
            DatabaseEvidence("Franklin", "token_required", "Franklin token required"),
            DatabaseEvidence("MTBP", "manual", "MTBP manual"),
            DatabaseEvidence("HSMD", "manual", "HSMD manual"),
        ]
    }

    ExcelReportWriter().write(result, output, evidence=evidence)

    workbook = openpyxl.load_workbook(output)
    assert workbook.sheetnames == ["With Artifacts", "Artifacts Removed"]
    ws = workbook["With Artifacts"]
    headers = [cell.value for cell in ws[1]]
    raw_headers = list(result.variants[0].raw)
    row = next(
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row[headers.index("Sample")] == variant.sample
    )
    workbook.close()

    assert headers[0] == "Skip Database Search (X)"
    assert headers[1 : len(raw_headers) + 1] == raw_headers
    assert headers[len(raw_headers) + 1 :] == [
        f"{database} Evidence"
        for database in [
            "ClinVar",
            "MTBP",
            "Franklin",
            "OncoKB",
            "COSMIC",
        ]
    ]
    assert row[headers.index("HGVSc")] == variant.raw["HGVSc"]
    assert "CIViC Evidence" not in headers


def test_database_selection_sheet_round_trips_x_marks(tmp_path):
    output = tmp_path / "review.xlsx"
    result = VariantProcessor().process(FIXTURE, "2026-07-26", output)
    skipped_key = DatabaseSearchService().variant_key(result.variants[1])

    ExcelReportWriter().write(
        result, output, database_skip_keys={skipped_key}
    )

    workbook = openpyxl.load_workbook(output)
    worksheet = workbook["With Artifacts"]
    assert worksheet["A1"].value == "Skip Database Search (X)"
    assert worksheet["A3"].value == "X"
    workbook.close()
    loaded = load_database_skip_keys(output)
    assert skipped_key in loaded
    assert DatabaseSearchService().variant_key(result.variants[0]) in loaded


def test_excel_export_writes_artifact_removed_sheet(tmp_path):
    output = tmp_path / "review.xlsx"
    result = VariantProcessor().process(FIXTURE, "2026-07-26", output)
    variant = result.variants[3]
    evidence = {
        DatabaseSearchService().variant_key(variant): [
            DatabaseEvidence("ClinVar", "found", "ClinVar summary"),
            DatabaseEvidence("gnomAD", "found", "gnomAD summary"),
            DatabaseEvidence("COSMIC", "found", "COSMIC summary"),
            DatabaseEvidence("CIViC", "found", "CIViC summary"),
            DatabaseEvidence("CancerMine", "found", "CancerMine summary"),
            DatabaseEvidence("DGIdb", "found", "DGIdb summary"),
            DatabaseEvidence("ClinGen Allele Registry", "found", "ClinGen summary"),
            DatabaseEvidence("cBioPortal", "found", "cBioPortal summary"),
            DatabaseEvidence("OncoKB", "token_required", "OncoKB token required"),
            DatabaseEvidence("Franklin", "unauthorized", "Franklin login was rejected"),
            DatabaseEvidence("MTBP", "manual", "MTBP manual"),
            DatabaseEvidence("HSMD", "manual", "HSMD manual"),
        ]
    }

    ExcelReportWriter().write(result, output, evidence=evidence)

    workbook = openpyxl.load_workbook(output)
    with_artifacts = workbook["With Artifacts"]
    artifacts_removed = workbook["Artifacts Removed"]
    with_headers = [cell.value for cell in with_artifacts[1]]
    with_samples = [
        row[with_headers.index("Sample")]
        for row in with_artifacts.iter_rows(min_row=2, values_only=True)
    ]
    removed_samples = [row[0] for row in artifacts_removed.iter_rows(min_row=2, values_only=True)]
    headers = [cell.value for cell in artifacts_removed[1]]
    row = next(row for row in artifacts_removed.iter_rows(min_row=2, values_only=True) if row[0] == variant.sample)
    workbook.close()

    assert "26OUM00001_VPM_S1_R1_001" in with_samples
    assert "26OUM00001_VPM_S1_R1_001" not in removed_samples
    assert row[headers.index("ClinVar Evidence")] == "[found] ClinVar summary"
    assert row[headers.index("Franklin Evidence")] == "[unauthorized] Franklin login was rejected"


def test_excel_export_keeps_row_coloring_on_raw_sheets(tmp_path):
    output = tmp_path / "review.xlsx"
    result = VariantProcessor().process(FIXTURE, "2026-07-26", output)
    by_patient = {variant.patient_id: variant for variant in result.variants}
    by_patient["26OUM00004"].history_matches = [{"Tier I": 2, "Tier II": 3}]
    by_patient["26OUM00005"].history_matches = [{"Germ": 10}]

    ExcelReportWriter().write(result, output)

    workbook = openpyxl.load_workbook(output)
    with_artifacts = workbook["With Artifacts"]
    artifacts_removed = workbook["Artifacts Removed"]
    by_sample = {
        row[1].value: row[1].fill.fgColor.rgb
        for row in with_artifacts.iter_rows(min_row=2)
    }
    removed_by_sample = {
        row[0].value: row[0].fill.fgColor.rgb
        for row in artifacts_removed.iter_rows(min_row=2)
    }
    workbook.close()

    assert by_sample["26OUM00001_VPM_S1_R1_001"] == "00FFC000"
    assert by_sample["26OUM00002_VPM_S2_R1_001"] == "00FFC000"
    assert by_sample["26OUM00005_VPM_S5_R1_001"] == "00FFC000"
    assert removed_by_sample["26OUM00004_VPM_S4_R1_001"] == "00FFFF00"
    assert "26OUM00005_VPM_S5_R1_001" not in removed_by_sample


def test_excel_review_layout_hides_reference_columns_and_keeps_evidence_compact(tmp_path):
    output = tmp_path / "review.xlsx"
    screenshot = tmp_path / "browser-evidence.png"
    screenshot.write_bytes(b"placeholder")
    result = VariantProcessor().process(FIXTURE, "2026-07-26", output)
    variant = result.variants[3]
    evidence = {
        DatabaseSearchService().variant_key(variant): [
            DatabaseEvidence(
                "OncoKB",
                "found",
                "oncogenic=Oncogenic",
                clinical_significance="Oncogenic",
                url="https://www.oncokb.org/example",
                raw={
                    "screenshot": str(screenshot),
                    "captured_at": "2026-08-01T12:00:00+00:00",
                },
            )
        ]
    }

    ExcelReportWriter().write(result, output, evidence=evidence)

    workbook = openpyxl.load_workbook(output)
    ws = workbook["With Artifacts"]
    headers = [cell.value for cell in ws[1]]
    evidence_column = headers.index("OncoKB Evidence") + 1
    evidence_cell = ws.cell(5, evidence_column)
    report_column = headers.index("Report") + 1
    assert ws.freeze_panes == "G2"
    assert ws.column_dimensions[openpyxl.utils.get_column_letter(report_column)].hidden
    assert all(not ws.row_dimensions[row].hidden for row in range(2, ws.max_row + 1))
    assert not evidence_cell.alignment.wrap_text
    assert ws.row_dimensions[5].height == 18
    workbook.close()
