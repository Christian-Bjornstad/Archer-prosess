from dataclasses import replace
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from PIL import Image

from archer_processor.core import DatabaseEvidence, VariantProcessor
from archer_processor.reports import PatientExcelReportWriter
from archer_processor.reports.patient_excel import IMAGE_DATABASES
from archer_processor.services import DatabaseSearchService


FIXTURE = Path(__file__).parent / "fixtures" / "sample_variants.tsv"


def test_patient_comment_and_long_hsmd_survive_regeneration(tmp_path):
    result = VariantProcessor().process(FIXTURE, "2026-09-05", tmp_path / "review.xlsx")
    variant = result.variants[3]
    output = tmp_path / "patient.xlsx"
    writer = PatientExcelReportWriter()
    writer.write_patient(result, variant.patient_id, [variant], output, {})
    workbook = openpyxl.load_workbook(output)
    sheet = workbook["Oversikt"]
    assert "E4:J7" in {str(area) for area in sheet.merged_cells.ranges}
    assert sheet["E4"].fill.fgColor.rgb == "00FFF3E8"
    sheet["E4"] = "Manuell vurdering\nBevares ved ny generering"
    sheet["D11"] = sheet["D11"].value.replace("HSMD -", "HSMD - " + "manuelt funn " * 20)
    workbook.save(output)
    workbook.close()
    writer.write_patient(result, variant.patient_id, [variant], output, {})
    workbook = openpyxl.load_workbook(output)
    sheet = workbook["Oversikt"]
    assert sheet["E4"].value == "Manuell vurdering\nBevares ved ny generering"
    assert "manuelt funn" in sheet["D11"].value
    assert sheet.row_dimensions[11].height > 160
    workbook.close()


def test_patient_overview_uses_light_blue_and_white_banding(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-08-11", tmp_path / "review.xlsx"
    )
    base = result.variants[3]
    strong = replace(base, raw={"Germ": 11}, af=0.35)
    weak = replace(
        base,
        source_row=base.source_row + 100,
        hgvsc="NM_015338.5:c.1935dup",
        raw={"Germ": 11},
        af=0.3499,
    )
    output = tmp_path / "patient.xlsx"

    PatientExcelReportWriter().write_patient(
        result, base.patient_id, [strong, weak], output, {}
    )

    workbook = openpyxl.load_workbook(output)
    try:
        overview = workbook["Oversikt"]
        # Light blue / white banding replaces the old green template.
        assert overview["A11"].fill.fgColor.rgb == "00EAF3FA"
        assert overview["A12"].fill.fgColor.rgb == "00FFFFFF"
    finally:
        workbook.close()


def test_patient_report_excludes_artifacts_but_keeps_them_in_data(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-09-03", tmp_path / "review.xlsx"
    )
    included = result.variants[3]
    artifact = replace(
        included,
        source_row=included.source_row + 100,
        hgvsc="NM_000546.6:c.525dup",
        raw={**included.raw, "HGVSc": "NM_000546.6:c.525dup"},
        matched_rules=["known_artifact"],
        decision="excluded",
    )
    result.variants = [included, artifact]
    output = tmp_path / "patient.xlsx"

    PatientExcelReportWriter().write_patient(
        result, included.patient_id, [included, artifact], output, {}
    )

    workbook = openpyxl.load_workbook(output)
    try:
        overview = workbook["Oversikt"]
        overview_hgvsc = [
            overview.cell(row, 2).value for row in range(11, overview.max_row + 1)
        ]
        assert included.hgvsc in overview_hgvsc
        assert artifact.hgvsc not in overview_hgvsc

        data = workbook["Data"]
        headers = [cell.value for cell in data[1]]
        hgvsc_column = headers.index("HGVSc") + 1
        data_hgvsc = [
            data.cell(row, hgvsc_column).value for row in range(2, data.max_row + 1)
        ]
        assert artifact.hgvsc in data_hgvsc

        report_text = " ".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            if sheet.title != "Data"
            for row in sheet.iter_rows()
            for cell in row
        )
        assert artifact.hgvsc not in report_text
    finally:
        workbook.close()


def test_patient_overview_sorts_variants_by_descending_af_with_missing_last(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-09-03", tmp_path / "review.xlsx"
    )
    base = result.variants[3]
    low = replace(base, source_row=101, hgvsc="NM_000546.6:c.100A>G", af=0.10)
    missing = replace(base, source_row=102, hgvsc="NM_000546.6:c.200A>G", af=None)
    high = replace(base, source_row=103, hgvsc="NM_000546.6:c.300A>G", af=0.25)
    output = tmp_path / "patient.xlsx"

    PatientExcelReportWriter().write_patient(
        result, base.patient_id, [low, missing, high], output, {}
    )

    workbook = openpyxl.load_workbook(output)
    try:
        overview = workbook["Oversikt"]
        assert [overview.cell(row, 2).value for row in range(11, 14)] == [
            high.hgvsc,
            low.hgvsc,
            missing.hgvsc,
        ]
    finally:
        workbook.close()


def test_patient_overview_row_height_follows_short_evidence_line_count(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-09-03", tmp_path / "review.xlsx"
    )
    variant = result.variants[3]
    output = tmp_path / "patient.xlsx"
    evidence = {
        f"{variant.sample}|{variant.hgvsc}": [
            DatabaseEvidence(
                "ClinVar",
                "found",
                "Pathogenic",
                clinical_significance="Pathogenic",
            )
        ]
    }

    PatientExcelReportWriter().write_patient(
        result, variant.patient_id, [variant], output, evidence
    )

    workbook = openpyxl.load_workbook(output)
    try:
        overview = workbook["Oversikt"]
        expected_lines = str(overview["D11"].value).count("\n") + 1
        assert expected_lines == 6
        assert overview.row_dimensions[11].height >= expected_lines * 16 + 12
    finally:
        workbook.close()


def test_patient_overview_preserves_manual_comment_and_hsmd_after_af_reordering(
    tmp_path,
):
    result = VariantProcessor().process(
        FIXTURE, "2026-09-03", tmp_path / "review.xlsx"
    )
    base = result.variants[3]
    first = replace(base, source_row=101, hgvsc="NM_000546.6:c.100A>G", af=0.30)
    second = replace(base, source_row=102, hgvsc="NM_000546.6:c.200A>G", af=0.10)
    output = tmp_path / "patient.xlsx"
    writer = PatientExcelReportWriter()

    writer.write_patient(result, base.patient_id, [first, second], output, {})
    workbook = openpyxl.load_workbook(output)
    try:
        overview = workbook["Oversikt"]
        headers = [overview.cell(10, column).value for column in range(1, 11)]
        assert headers == [
            "Gen",
            "HGVSc",
            "HGVSp",
            "Kort evidens",
            "Kommentar",
            "MTBP",
            "Franklin",
            "ClinVar",
            "OncoKB",
            "COSMIC",
        ]
        second_row = next(
            row
            for row in range(11, overview.max_row + 1)
            if overview.cell(row, 2).value == second.hgvsc
        )
        overview.cell(second_row, 4, "MTBP - gammelt\nHSMD - intern klassifikasjon")
        overview.cell(second_row, 5, "Vurdert manuelt")
        workbook.save(output)
    finally:
        workbook.close()

    first.af = 0.05
    second.af = 0.40
    writer.write_patient(result, base.patient_id, [first, second], output, {})

    regenerated = openpyxl.load_workbook(output)
    try:
        overview = regenerated["Oversikt"]
        second_row = next(
            row
            for row in range(11, overview.max_row + 1)
            if overview.cell(row, 2).value == second.hgvsc
        )
        assert second_row == 11
        assert overview.cell(second_row, 5).value == "Vurdert manuelt"
        assert "HSMD - intern klassifikasjon" in overview.cell(second_row, 4).value
        assert "MTBP - gammelt" not in overview.cell(second_row, 4).value
    finally:
        regenerated.close()


def test_patient_attachment_contains_combined_mtbp_report_only_once(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-09-03", tmp_path / "review.xlsx"
    )
    first = result.variants[3]
    second = replace(
        first,
        source_row=first.source_row + 1,
        hgvsc="NM_000546.6:c.743G>A",
        hgvsp="p.R248Q",
    )
    screenshot = tmp_path / "mtbp-full-report.png"
    Image.new("RGB", (1200, 800), "white").save(screenshot)
    evidence = {
        DatabaseSearchService().variant_key(variant): [
            DatabaseEvidence(
                "MTBP",
                "found",
                "matched",
                raw={"patient_report_screenshot": str(screenshot)},
            )
        ]
        for variant in (first, second)
    }
    output = tmp_path / "patient.xlsx"

    PatientExcelReportWriter().write_patient(
        result, first.patient_id, [first, second], output, evidence
    )

    workbook = openpyxl.load_workbook(output)
    try:
        attachment = workbook["Vedlegg"]
        assert len(attachment._images) == 1
        assert attachment["A3"].value == "MTBP – samlet pasientrapport"
    finally:
        workbook.close()


def test_patient_data_sheet_includes_artifacts_without_skip_column(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-08-11", tmp_path / "review.xlsx"
    )
    base = result.variants[3]
    strong = replace(base, raw={**base.raw, "Germ": 11}, af=0.35)
    artifact_raw = {
        **base.raw,
        "HGVSc": "NM_000546.6:c.525dup",
        "Tier I": 0,
        "Tier II": 0,
    }
    artifact = replace(
        base,
        source_row=base.source_row + 100,
        hgvsc="NM_000546.6:c.525dup",
        raw=artifact_raw,
        matched_rules=["known-artifact"],
    )
    result.variants = [strong, artifact]
    output = tmp_path / "patient.xlsx"

    PatientExcelReportWriter().write_patient(
        result, base.patient_id, [strong], output, {}
    )

    workbook = openpyxl.load_workbook(output)
    try:
        data = workbook["Data"]
        headers = [cell.value for cell in data[1]]
        assert headers[0] == "Sample"
        assert "Skip Database Search (X)" not in headers
        assert headers[-5:] == [
            "ClinVar Evidence",
            "MTBP Evidence",
            "Franklin Evidence",
            "OncoKB Evidence",
            "COSMIC Evidence",
        ]
        assert data.max_row == 3
        assert data.sheet_properties.tabColor.rgb == "004F8A5B"
        hgvsc_column = headers.index("HGVSc") + 1
        colors_by_hgvsc = {
            data.cell(row, hgvsc_column).value: data.cell(row, 1).fill.fgColor.rgb
            for row in range(2, data.max_row + 1)
        }
        assert colors_by_hgvsc[strong.hgvsc] == "00C6EFCE"
        assert colors_by_hgvsc[artifact.hgvsc] == "00FFC000"
        report_column = headers.index("Report") + 1
        assert data.column_dimensions[
            openpyxl.utils.get_column_letter(report_column)
        ].hidden
    finally:
        workbook.close()


def test_patient_excel_report_uses_requested_sheet_layout_and_image_order(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-08-03", tmp_path / "review.xlsx"
    )
    variant = result.variants[3]
    image_paths = []
    for name, size, color in [
        ("mtbp.png", (1200, 600), "#FFF3E8"),
        ("franklin-full.png", (1400, 2000), "#D9EAF7"),
        ("franklin-assessment.png", (1400, 500), "#EAF3FA"),
        ("clinvar.png", (1200, 450), "#F3F6F8"),
        ("oncokb.png", (1400, 700), "#EAF5ED"),
        ("cosmic.png", (1400, 700), "#EAF3FA"),
    ]:
        path = tmp_path / name
        Image.new("RGB", size, color).save(path)
        image_paths.append(path)

    evidence = {
        DatabaseSearchService().variant_key(variant): [
            DatabaseEvidence(
                "ClinVar",
                "found",
                "Pathogenic with expert-panel review.",
                accession="VCV000012345.1",
                clinical_significance="Pathogenic",
                url="https://www.ncbi.nlm.nih.gov/clinvar/variation/12345/",
                raw={
                    "screenshots": [{
                        "label": "Classification summary",
                        "path": str(image_paths[3]),
                        "url": "https://www.ncbi.nlm.nih.gov/clinvar/variation/12345/",
                    }]
                },
            ),
            DatabaseEvidence(
                "COSMIC",
                "found",
                "COSMIC public dataset: 2 records returned; primary_sites=haematopoietic_and_lymphoid_tissue",
                accession="20479064",
                url="https://cancer.sanger.ac.uk/cosmic/search?q=TP53",
                raw={
                    "query": "TP53 c.524G>A",
                    "records": [
                        {
                            "GeneName": "TP53",
                            "MutationCDS": "c.524G>A",
                            "MutationAA": "p.R175H",
                            "MutationID": "20479064",
                            "LegacyMutationID": "COSM99023",
                            "GenomicMutationID": "COSV52661038",
                            "PrimarySite": "haematopoietic_and_lymphoid_tissue",
                            "PrimaryHistology": "haematopoietic_neoplasm",
                            "PubmedPMID": "12345; 67890",
                            "GRChVer": "37",
                        },
                        {
                            "GeneName": "TP53",
                            "MutationCDS": "c.524G>A",
                            "MutationAA": "p.R175H",
                            "MutationID": "20479064",
                            "PrimarySite": "lung",
                            "PrimaryHistology": "carcinoma",
                            "PubmedPMID": "24680",
                            "GRChVer": "37",
                        },
                    ],
                    "screenshots": [{
                        "label": "Overview",
                        "path": str(image_paths[5]),
                        "url": "https://cancer.sanger.ac.uk/cosmic/example",
                    }],
                },
            ),
            DatabaseEvidence(
                "Franklin",
                "found",
                "classification=Pathogenic",
                clinical_significance="Pathogenic",
                url="https://franklin.genoox.com/example",
                raw={
                    "screenshots": [
                        {
                            "label": "Full computed-classification page",
                            "path": str(image_paths[1]),
                            "url": "https://franklin.genoox.com/example",
                        },
                        {
                            "label": "Predictions and population frequencies",
                            "path": str(image_paths[2]),
                            "url": "https://franklin.genoox.com/example?app=assessment-tools",
                        },
                    ]
                },
            ),
            DatabaseEvidence(
                "OncoKB",
                "found",
                "oncogenic=Oncogenic",
                url="https://www.oncokb.org/example",
                raw={
                    "screenshots": [
                        {
                            "label": "Variant overview and mutation effect",
                            "path": str(image_paths[4]),
                            "url": "https://www.oncokb.org/example",
                        }
                    ]
                },
            ),
            DatabaseEvidence(
                "MTBP",
                "found",
                "Putative functionally relevant variant",
                url="https://mtbp.org/example",
                raw={
                    "screenshots": [
                        {
                            "label": "Alteration-centric functional evidence",
                            "path": str(image_paths[0]),
                            "url": "https://mtbp.org/example",
                        }
                    ]
                },
            ),
        ]
    }
    output = tmp_path / "patient.xlsx"

    PatientExcelReportWriter().write_patient(
        result,
        variant.patient_id,
        [variant],
        output,
        evidence,
    )

    workbook = openpyxl.load_workbook(output)
    assert workbook.sheetnames == ["Oversikt", "Vedlegg", "Data", "TP53"]
    overview = workbook["Oversikt"]
    assert overview["A1"].value == "VPM-tolkning – 26OUM00004"
    assert overview["A3"].value is None
    assert [overview.cell(10, column).value for column in range(1, 11)] == [
        "Gen", "HGVSc", "HGVSp", "Kort evidens",
        "Kommentar", "MTBP", "Franklin", "ClinVar", "OncoKB", "COSMIC",
    ]
    assert "ClinVar - Pathogenic" in overview["D11"].value
    assert "HSMD -" in overview["D11"].value
    assert overview["H11"].value == "Pathogenic"
    assert overview["H11"].hyperlink.target.endswith("/12345/")
    assert overview["F11"].hyperlink is None
    assert workbook["Vedlegg"]["A1"].value == "26OUM00004"
    variant_sheet = workbook["TP53"]
    assert variant_sheet["A6"].hyperlink is None
    assert variant_sheet["A7"].hyperlink is not None
    screenshot_heading_row = next(
        row
        for row in range(1, variant_sheet.max_row + 1)
        if variant_sheet.cell(row, 1).value == "Skjermbilder"
    )
    assert all(
        variant_sheet.cell(row, 1).hyperlink is None
        for row in range(screenshot_heading_row + 1, variant_sheet.max_row + 1)
    )
    assert len(variant_sheet._images) == 6
    assert IMAGE_DATABASES == ("MTBP", "Franklin", "ClinVar", "OncoKB", "COSMIC")
    workbook.close()


def test_patient_excel_uses_variant_detail_only_for_duplicate_gene(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-08-03", tmp_path / "review.xlsx"
    )
    first = result.variants[3]
    second = replace(
        first,
        source_row=first.source_row + 1,
        hgvsc="NM_000546.6:c.743G>A",
        hgvsp="",
    )
    output = tmp_path / "duplicate.xlsx"

    PatientExcelReportWriter().write_patient(
        result, first.patient_id, [first, second], output, {}
    )

    workbook = openpyxl.load_workbook(output)
    assert workbook.sheetnames == [
        "Oversikt",
        "Vedlegg",
        "Data",
        "TP53 p.R175H",
        "TP53 c.743G>A",
    ]
    workbook.close()


def test_patient_excel_filename_is_dit_vpm_tolkning(tmp_path):
    result = VariantProcessor().process(
        FIXTURE, "2026-08-03", tmp_path / "review.xlsx"
    )

    outputs = PatientExcelReportWriter().write_all(result, tmp_path / "patients", {})

    assert outputs[0].name.endswith("_VPM_Tolkning_APP.xlsx")


def test_write_patient_saves_via_temporary_file_then_replaces(tmp_path, monkeypatch):
    result = VariantProcessor().process(
        FIXTURE, "2026-09-03", tmp_path / "review.xlsx"
    )
    variant = result.variants[0]
    output = tmp_path / "VEDLEGG_APP" / f"{variant.patient_id}_VPM_Tolkning_APP.xlsx"
    saved_paths = []
    real_save = Workbook.save

    def recording_save(workbook_self, filename):
        saved_paths.append(Path(filename))
        return real_save(workbook_self, filename)

    monkeypatch.setattr(Workbook, "save", recording_save)

    PatientExcelReportWriter().write_patient(
        result, variant.patient_id, [variant], output, {}
    )

    assert len(saved_paths) == 1
    assert saved_paths[0] != output
    assert saved_paths[0].name.endswith(".tmp.xlsx")
    assert saved_paths[0].parent == output.parent
    assert output.exists()
    assert [entry for entry in output.parent.iterdir()] == [output]


def test_failed_save_keeps_existing_workbook_and_removes_temporary_file(
    tmp_path, monkeypatch
):
    result = VariantProcessor().process(
        FIXTURE, "2026-09-03", tmp_path / "review.xlsx"
    )
    variant = result.variants[0]
    output = tmp_path / f"{variant.patient_id}_VPM_Tolkning_APP.xlsx"
    writer = PatientExcelReportWriter()
    writer.write_patient(result, variant.patient_id, [variant], output, {})
    before = output.read_bytes()

    def exploding_save(workbook_self, filename):
        Path(filename).write_bytes(b"partial bytes")
        raise RuntimeError("save exploded")

    monkeypatch.setattr(Workbook, "save", exploding_save)

    try:
        writer.write_patient(result, variant.patient_id, [variant], output, {})
    except RuntimeError:
        pass

    assert output.read_bytes() == before
    assert [entry for entry in output.parent.iterdir()] == [output]
