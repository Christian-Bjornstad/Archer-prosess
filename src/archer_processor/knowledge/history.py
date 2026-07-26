from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from archer_processor.core.models import VariantRecord


class VariantHistoryRepository:
    wanted_columns = {
        "Sample",
        "Classification",
        "Symbol",
        "HGVSp",
        "HGVSc",
        "Depth",
        "AO",
        "AF",
        "Type",
        "Quality Score",
        "TO",
        "Rept",
        "Tier I",
        "Tier II",
        "Tier III",
        "Tier IV",
        "Germ",
        "Artf",
    }

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self._loaded = False
        self._by_hgvsc: dict[str, list[dict[str, Any]]] = {}

    def load(self) -> None:
        if self._loaded:
            return
        if not self.workbook_path.exists():
            self._loaded = True
            return
        workbook = openpyxl.load_workbook(self.workbook_path, read_only=True, data_only=True)
        worksheet = workbook["RESULTAT"] if "RESULTAT" in workbook.sheetnames else workbook.active
        header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        indexes = {
            name: header.index(name)
            for name in self.wanted_columns
            if name in header
        }
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {name: values[index] if index < len(values) else None for name, index in indexes.items()}
            hgvsc = str(row.get("HGVSc") or "").strip()
            if hgvsc:
                self._by_hgvsc.setdefault(hgvsc, []).append(row)
        workbook.close()
        self._loaded = True

    def find(self, hgvsc: str, limit: int = 8) -> list[dict[str, Any]]:
        self.load()
        return self._by_hgvsc.get(hgvsc, [])[:limit]

    def annotate(self, variants: list[VariantRecord]) -> None:
        for variant in variants:
            variant.history_matches = self.find(variant.hgvsc)

    def stats(self) -> dict[str, int]:
        self.load()
        unique_samples = {
            str(entry.get("Sample") or "")
            for entries in self._by_hgvsc.values()
            for entry in entries
        }
        return {
            "unique_hgvsc": len(self._by_hgvsc),
            "total_entries": sum(len(entries) for entries in self._by_hgvsc.values()),
            "unique_samples": len(unique_samples),
        }
