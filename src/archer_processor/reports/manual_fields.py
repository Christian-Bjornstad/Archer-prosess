from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

from archer_processor.core.models import VariantRecord


@dataclass(frozen=True, slots=True)
class ManualVariantFields:
    comment: str = ""
    hsmd: str = "HSMD -"


def variant_manual_key(patient_id: str, variant: VariantRecord) -> str:
    return _manual_key(patient_id, variant.symbol, variant.hgvsc, variant.hgvsp)


def read_manual_fields(
    path: Path, patient_id: str
) -> dict[str, ManualVariantFields]:
    if not path.is_file():
        return {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        if "Oversikt" not in workbook.sheetnames:
            return {}
        sheet = workbook["Oversikt"]
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[10]
            if cell.value not in {None, ""}
        }
        required = {"Gen", "HGVSc", "HGVSp", "Kort evidens"}
        if not required.issubset(headers):
            return {}
        comment_column = headers.get("Kommentar")
        preserved: dict[str, ManualVariantFields] = {}
        for row in range(11, sheet.max_row + 1):
            gene = str(sheet.cell(row, headers["Gen"]).value or "").strip()
            hgvsc = str(sheet.cell(row, headers["HGVSc"]).value or "").strip()
            hgvsp = str(sheet.cell(row, headers["HGVSp"]).value or "").strip()
            if not gene and not hgvsc:
                continue
            compact = str(
                sheet.cell(row, headers["Kort evidens"]).value or ""
            )
            hsmd = next(
                (
                    line.strip()
                    for line in compact.splitlines()
                    if line.strip().casefold().startswith("hsmd")
                ),
                "HSMD -",
            )
            comment = (
                str(sheet.cell(row, comment_column).value or "").strip()
                if comment_column is not None
                else ""
            )
            preserved[_manual_key(patient_id, gene, hgvsc, hgvsp)] = (
                ManualVariantFields(comment=comment, hsmd=hsmd)
            )
        return preserved
    finally:
        workbook.close()


def _manual_key(patient_id: str, gene: str, hgvsc: str, hgvsp: str) -> str:
    return "|".join(
        [
            patient_id.strip().casefold(),
            gene.strip().casefold(),
            hgvsc.strip().casefold(),
            hgvsp.strip().casefold(),
        ]
    )
