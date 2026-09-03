from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage

from archer_processor.core.highlights import variant_highlight
from archer_processor.core.models import DatabaseEvidence, ProcessingResult, VariantRecord
from archer_processor.core.sorting import variant_sort_key
from archer_processor.reports.excel_report import ExcelReportWriter


REPORT_DATABASES = ("MTBP", "Franklin", "ClinVar", "OncoKB", "COSMIC")
TEXT_DATABASES = REPORT_DATABASES
IMAGE_DATABASES = REPORT_DATABASES
COSMIC_FIELDS = (
    "AccessionNumber",
    "GeneCDS_Length",
    "GeneName",
    "HGNC_ID",
    "MutationAA",
    "MutationCDS",
    "MutationDescription",
    "MutationGenomePosition",
    "MutationStrand",
    "MutationID",
    "LegacyMutationID",
    "GenomicMutationID",
    "Name",
    "PrimaryHistology",
    "PrimarySite",
    "PubmedPMID",
    "Site",
    "GRChVer",
    "COSMIC_GENE_ID",
    "COSMIC_PHENOTYPE_ID",
)
COSMIC_FIELD_DESCRIPTIONS = {
    "AccessionNumber": "Transcript accession used by COSMIC for the mutation annotation.",
    "GeneCDS_Length": "Length of the gene coding sequence in bases.",
    "GeneName": "HGNC gene symbol recorded by COSMIC.",
    "HGNC_ID": "HGNC identifier for the gene.",
    "MutationAA": "Protein-level amino-acid change.",
    "MutationCDS": "Coding-DNA change.",
    "MutationDescription": "COSMIC mutation consequence or description.",
    "MutationGenomePosition": "Genomic coordinate or interval recorded by COSMIC.",
    "MutationStrand": "Transcript/genomic strand annotation.",
    "MutationID": "Current COSMIC mutation identifier.",
    "LegacyMutationID": "Legacy COSMIC mutation identifier retained for cross-reference.",
    "GenomicMutationID": "COSMIC genomic mutation identifier.",
    "Name": "Combined transcript, coding, and protein mutation name.",
    "PrimaryHistology": "Primary tumour histology associated with the record.",
    "PrimarySite": "Primary anatomical tumour site associated with the record.",
    "PubmedPMID": "PubMed identifier for the supporting publication.",
    "Site": "Combined site and histology label supplied by the public dataset.",
    "GRChVer": "Genome assembly version used for the genomic coordinates.",
    "COSMIC_GENE_ID": "COSMIC gene identifier.",
    "COSMIC_PHENOTYPE_ID": "COSMIC phenotype identifier for the tumour context.",
}
COSMIC_PUBLIC_API_URL = "https://clinicaltables.nlm.nih.gov/apidoc/cosmic/v4/doc.html"


class PatientExcelReportWriter:
    """Create one image-led evidence workbook per patient/DIT identifier."""

    colors = {
        "navy": "163B5C",
        "blue": "2F75B5",
        "pale_blue": "EAF3FA",
        "pale_green": "EAF5ED",
        "green": "4F8A5B",
        "strong_green": "C6EFCE",
        "weak_green": "E9F6EF",
        "orange": "FFC000",
        "light_orange": "F4B183",
        "pale_orange": "FFF3E8",
        "gray": "F3F6F8",
        "muted": "5E6A73",
        "border": "CBD7E1",
        "white": "FFFFFF",
    }

    def write_all(
        self,
        result: ProcessingResult,
        output_directory: Path,
        evidence: dict[str, list[DatabaseEvidence]] | None = None,
        variants: list[VariantRecord] | None = None,
    ) -> list[Path]:
        evidence = evidence or {}
        grouped: dict[str, list[VariantRecord]] = defaultdict(list)
        report_variants = result.included if variants is None else variants
        for variant in report_variants:
            grouped[variant.patient_id].append(variant)
        output_directory.mkdir(parents=True, exist_ok=True)
        outputs: list[Path] = []
        for patient_id, variants in sorted(grouped.items()):
            safe_patient = re.sub(r"[^A-Za-z0-9_-]+", "_", patient_id).strip("_")
            output = output_directory / f"{safe_patient}_Myolid_Tolkning_APP.xlsx"
            self.write_patient(result, patient_id, variants, output, evidence)
            outputs.append(output)
        return outputs

    def write_patient(
        self,
        result: ProcessingResult,
        patient_id: str,
        variants: list[VariantRecord],
        output_path: Path,
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> Path:
        variants = sorted(variants, key=variant_sort_key)
        workbook = Workbook()
        placeholder = workbook.active
        self._overview_sheet(workbook, result, patient_id, variants, evidence)
        workbook.remove(placeholder)
        self._attachment_sheet(workbook, patient_id)
        patient_data = [
            variant for variant in result.variants
            if variant.patient_id == patient_id
        ]
        self._data_sheet(workbook, patient_data, evidence)
        gene_counts = Counter((variant.symbol or "Variant").casefold() for variant in variants)
        used_names = {"Oversikt", "Vedlegg", "Data"}
        for index, variant in enumerate(variants, start=1):
            title = self._variant_sheet_name(
                variant,
                index,
                used_names,
                duplicate_gene=gene_counts[(variant.symbol or "Variant").casefold()] > 1,
            )
            used_names.add(title)
            self._variant_sheet(
                workbook,
                title,
                variant,
                evidence.get(self._key(variant), []),
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _report_sheet(
        self,
        ws,
        patient_id: str,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> None:
        self._base_sheet(ws)
        ws.sheet_properties.tabColor = self.colors["navy"]
        ws.merge_cells("A1:L1")
        ws["A1"] = f"Patient ID: {patient_id}"
        ws["A1"].font = Font(size=18, bold=True, color=self.colors["white"])
        ws["A1"].fill = PatternFill("solid", fgColor=self.colors["navy"])
        ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[1].height = 32

        row = 3
        for index, variant in enumerate(variants, start=1):
            if index > 1:
                row += 1
            by_database = self._by_database(evidence.get(self._key(variant), []))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            ws.cell(row, 1, self._variant_heading(variant))
            self._section_style(ws.cell(row, 1))
            ws.row_dimensions[row].height = 26
            row += 1

            for database in TEXT_DATABASES:
                items = by_database.get(database, [])
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
                label = ws.cell(row, 1, database)
                value = ws.cell(row, 3, self._text_evidence(items))
                label.fill = PatternFill("solid", fgColor=self.colors["pale_blue"])
                label.font = Font(bold=True, color=self.colors["navy"])
                label.alignment = Alignment(vertical="top")
                value.fill = PatternFill("solid", fgColor=self.colors["gray"])
                value.alignment = Alignment(vertical="top", wrap_text=True)
                for cell in ws[row]:
                    cell.border = self._border()
                if items and items[0].url and database != "MTBP":
                    label.hyperlink = items[0].url
                    label.style = "Hyperlink"
                ws.row_dimensions[row].height = 82 if database == "COSMIC" else 58
                row += 1

            row += 1
            for database in IMAGE_DATABASES:
                items = by_database.get(database, [])
                records = self._screenshot_records(items)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
                ws.cell(row, 1, database)
                self._image_section_style(ws.cell(row, 1), database)
                ws.row_dimensions[row].height = 24
                row += 1
                if not records:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
                    ws.cell(row, 1, self._text_evidence(items) or "No screenshot available.")
                    ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row, 1).fill = PatternFill(
                        "solid", fgColor=self.colors["pale_orange"]
                    )
                    ws.row_dimensions[row].height = 36
                    row += 2
                    continue
                for record in records:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
                    caption = ws.cell(row, 1, record["label"])
                    caption.font = Font(bold=True, color=self.colors["navy"])
                    row += 1
                    row = self._add_image(ws, Path(record["path"]), row)
                    row += 1

        ws.freeze_panes = "A2"
        ws.print_area = f"A1:L{max(3, row)}"

    @staticmethod
    def _variant_heading(variant: VariantRecord) -> str:
        return " | ".join(
            part
            for part in [variant.symbol, variant.hgvsc, variant.hgvsp]
            if part
        )

    def _overview_sheet(
        self,
        workbook: Workbook,
        result: ProcessingResult,
        patient_id: str,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> None:
        ws = workbook.create_sheet("Oversikt")
        self._base_sheet(ws)
        ws.sheet_properties.tabColor = self.colors["navy"]
        ws.merge_cells("A1:I2")
        ws["A1"] = f"VPM-tolkning – {patient_id}"
        self._title_style(ws["A1"])
        self._info_row(ws, 5, "DIT/pasientnummer", patient_id, end_column=9)
        self._info_row(ws, 6, "Rapportdato", result.run_date, end_column=9)
        self._info_row(ws, 7, "Antall varianter", len(variants), end_column=9)

        ws.merge_cells("A9:I9")
        ws["A9"] = "Varianter og signifikant evidens"
        self._section_style(ws["A9"])
        headers = ["Gen", "HGVSc", "HGVSp", "Kort evidens", *REPORT_DATABASES]
        for column, header in enumerate(headers, start=1):
            cell = ws.cell(10, column, header)
            self._header_style(cell)
        for row, variant in enumerate(variants, start=11):
            by_database = self._by_database(evidence.get(self._key(variant), []))
            values: list[Any] = [
                variant.symbol,
                variant.hgvsc,
                variant.hgvsp,
                "\n".join(
                    f"{database} - {self._compact_evidence(by_database.get(database, []))}"
                    for database in REPORT_DATABASES
                ),
                *[
                    self._compact_evidence(by_database.get(database, []))
                    for database in REPORT_DATABASES
                ],
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row, column, value)
                cell.border = self._border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for offset, database in enumerate(REPORT_DATABASES, start=5):
                items = by_database.get(database, [])
                if items and items[0].url and database != "MTBP":
                    ws.cell(row, offset).hyperlink = items[0].url
                    ws.cell(row, offset).style = "Hyperlink"
                    ws.cell(row, offset).alignment = Alignment(
                        vertical="top", wrap_text=True
                    )
            # Light blue / white theme for the overview table; artifact rows
            # keep their orange attention color.
            fill_color = {
                "artifact": self.colors["orange"],
                "artifact_light": self.colors["light_orange"],
                "tier": self.colors["pale_blue"],
                "germline": self.colors["pale_blue"],
                "germline_low_af": self.colors["white"],
            }.get(variant_highlight(variant))
            if not fill_color:
                fill_color = (
                    self.colors["white"]
                    if (row - 11) % 2
                    else self.colors["pale_blue"]
                )
            if fill_color:
                for column in range(1, 10):
                    ws.cell(row, column).fill = PatternFill(
                        "solid", fgColor=fill_color
                    )
            ws.row_dimensions[row].height = 76
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 31
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 52
        for column in "EFGHI":
            ws.column_dimensions[column].width = 22
        ws.auto_filter.ref = f"A10:I{max(10, 10 + len(variants))}"
        ws.freeze_panes = "A10"
        ws.print_area = f"A1:I{max(16, 11 + len(variants))}"

    def _attachment_sheet(self, workbook: Workbook, patient_id: str) -> None:
        ws = workbook.create_sheet("Vedlegg")
        self._base_sheet(ws)
        ws.sheet_properties.tabColor = self.colors["muted"]
        ws["A1"] = patient_id
        ws["A1"].font = Font(size=18, bold=True, color=self.colors["navy"])
        ws.column_dimensions["A"].width = max(18, len(patient_id) + 4)

    def _data_sheet(
        self,
        workbook: Workbook,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> None:
        raw_writer = ExcelReportWriter()
        raw_writer._raw_variant_sheet(
            workbook,
            "Data",
            variants,
            evidence,
            include_selection=False,
        )
        ws = workbook["Data"]
        ws.sheet_properties.tabColor = self.colors["green"]
        ws.sheet_view.showGridLines = False

    def _variant_sheet(
        self,
        workbook: Workbook,
        title: str,
        variant: VariantRecord,
        evidence_items: list[DatabaseEvidence],
    ) -> None:
        ws = workbook.create_sheet(title)
        self._base_sheet(ws)
        by_database = self._by_database(evidence_items)
        ws.merge_cells("A1:L2")
        ws["A1"] = self._variant_heading(variant)
        self._title_style(ws["A1"])
        ws.merge_cells("A3:L3")
        ws["A3"] = " | ".join(
            part
            for part in [
                variant.hgvsp,
                variant.genomic_location,
                f"AF {variant.af:.2%}" if variant.af is not None else "",
                f"Depth {variant.depth}" if variant.depth is not None else "",
                variant.consequence,
            ]
            if part
        )
        ws["A3"].font = Font(color=self.colors["muted"], italic=True)
        ws["A3"].alignment = Alignment(wrap_text=True)

        row = 5
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        ws.cell(row, 1, "Kort evidensoversikt og kildelenker")
        self._section_style(ws.cell(row, 1))
        row += 1
        for database in REPORT_DATABASES:
            items = by_database.get(database, [])
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
            label = ws.cell(row, 1, database)
            value = ws.cell(row, 3, self._compact_evidence(items))
            label.fill = PatternFill("solid", fgColor=self.colors["pale_blue"])
            label.font = Font(bold=True, color=self.colors["navy"])
            label.alignment = Alignment(vertical="top")
            value.fill = PatternFill("solid", fgColor=self.colors["gray"])
            value.alignment = Alignment(vertical="top", wrap_text=True)
            for cell in ws[row]:
                cell.border = self._border()
            if items and items[0].url and database != "MTBP":
                label.hyperlink = items[0].url
                label.style = "Hyperlink"
            if items and items[0].url and database != "MTBP":
                value.hyperlink = items[0].url
                value.style = "Hyperlink"
                value.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 34
            row += 1

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        ws.cell(row, 1, "Skjermbilder")
        self._section_style(ws.cell(row, 1))
        row += 1
        for database in IMAGE_DATABASES:
            items = by_database.get(database, [])
            records = self._screenshot_records(items)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            ws.cell(row, 1, database)
            self._image_section_style(ws.cell(row, 1), database)
            row += 1
            if not records:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
                ws.cell(row, 1, "Skjermbilde ikke tilgjengelig.")
                ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.colors["pale_orange"])
                ws.row_dimensions[row].height = 42
                row += 2
                continue
            for record in records:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
                caption = ws.cell(row, 1, record["label"])
                caption.font = Font(bold=True, color=self.colors["navy"])
                caption.fill = PatternFill("solid", fgColor=self.colors["pale_blue"])
                row += 1
                row = self._add_image(ws, Path(record["path"]), row)
                row += 1
        ws.freeze_panes = "A5"
        ws.print_area = f"A1:L{max(12, row)}"

    def _evidence_data_sheet(
        self,
        workbook: Workbook,
        patient_id: str,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> None:
        ws = workbook.create_sheet("Evidence Data")
        self._base_sheet(ws)
        headers = [
            "Patient",
            "Gene",
            "HGVSc",
            "HGVSp",
            "Database",
            "Status",
            "Clinical significance",
            "Accession",
            "Summary",
            "Source URL",
            "Captured at",
        ]
        self._write_headers(ws, headers)
        row = 2
        for variant in variants:
            for item in evidence.get(self._key(variant), []):
                if item.database not in REPORT_DATABASES:
                    continue
                values = [
                    patient_id,
                    variant.symbol,
                    variant.hgvsc,
                    variant.hgvsp,
                    item.database,
                    item.status,
                    item.clinical_significance,
                    item.accession,
                    item.summary,
                    "" if item.database == "MTBP" else item.url,
                    item.raw.get("captured_at", ""),
                ]
                for column, value in enumerate(values, start=1):
                    cell = ws.cell(row, column, value)
                    cell.border = self._border()
                    cell.alignment = Alignment(vertical="top", wrap_text=column in {7, 9, 10})
                if item.url and item.database != "MTBP":
                    ws.cell(row, 10).hyperlink = item.url
                    ws.cell(row, 10).style = "Hyperlink"
                ws.row_dimensions[row].height = 48
                row += 1
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:K{max(1, row - 1)}"
        self._fit_columns(ws, max_width=65)

    def _cosmic_records_sheet(
        self,
        workbook: Workbook,
        patient_id: str,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> None:
        ws = workbook.create_sheet("COSMIC Records")
        self._base_sheet(ws)
        headers = ["Patient", "Variant Gene", "Variant HGVSc", "Query", *COSMIC_FIELDS]
        self._write_headers(ws, headers)
        row = 2
        for variant in variants:
            for item in evidence.get(self._key(variant), []):
                if item.database != "COSMIC":
                    continue
                for record in item.raw.get("records") or []:
                    values = [
                        patient_id,
                        variant.symbol,
                        variant.hgvsc,
                        item.raw.get("query", ""),
                        *[self._cell_value(record.get(field)) for field in COSMIC_FIELDS],
                    ]
                    for column, value in enumerate(values, start=1):
                        cell = ws.cell(row, column, value)
                        cell.border = self._border()
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                    row += 1
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, row - 1)}"
        self._fit_columns(ws, max_width=42)

    def _cosmic_field_guide_sheet(self, workbook: Workbook) -> None:
        ws = workbook.create_sheet("COSMIC Field Guide")
        self._base_sheet(ws)
        ws.merge_cells("A1:F2")
        ws["A1"] = "COSMIC Public Dataset Field Guide"
        self._title_style(ws["A1"])
        ws.merge_cells("A3:F3")
        ws["A3"] = (
            "Fields preserved from the NLM Clinical Tables COSMIC v4 endpoint. "
            "The COSMIC Records sheet contains the unfiltered returned values."
        )
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
        ws["A3"].font = Font(italic=True, color=self.colors["muted"])
        ws.merge_cells("A4:F4")
        ws["A4"] = "Open the official NLM COSMIC v4 field documentation"
        ws["A4"].hyperlink = COSMIC_PUBLIC_API_URL
        ws["A4"].style = "Hyperlink"
        headers = ["Field", "Meaning", "Report use"]
        for column, header in enumerate(headers, start=1):
            cell = ws.cell(5, column, header)
            self._header_style(cell)
        for row, field in enumerate(COSMIC_FIELDS, start=6):
            values = [
                field,
                COSMIC_FIELD_DESCRIPTIONS[field],
                "Available for review; retain or remove after clinical-team assessment.",
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row, column, value)
                cell.border = self._border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row, 1).font = Font(bold=True, color=self.colors["navy"])
            ws.row_dimensions[row].height = 44
        widths = {"A": 27, "B": 60, "C": 48, "D": 3, "E": 3, "F": 3}
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A5:C{5 + len(COSMIC_FIELDS)}"
        ws.print_area = f"A1:C{5 + len(COSMIC_FIELDS)}"

    def _add_image(self, ws, path: Path, row: int) -> int:
        if not path.is_file():
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            ws.cell(row, 1, f"Screenshot file not found: {path}")
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.colors["pale_orange"])
            return row + 1
        with PillowImage.open(path) as source:
            width, height = source.size
        scale = min(1.0, 1120 / max(1, width), 2200 / max(1, height))
        image = ExcelImage(str(path))
        image.width = max(1, int(width * scale))
        image.height = max(1, int(height * scale))
        ws.add_image(image, f"A{row}")
        rows_needed = max(1, math.ceil(image.height / 38))
        for image_row in range(row, row + rows_needed):
            ws.row_dimensions[image_row].height = 28.5
        return row + rows_needed

    def _base_sheet(self, ws) -> None:
        ws.sheet_view.showGridLines = False
        for column in range(1, 13):
            ws.column_dimensions[get_column_letter(column)].width = 13
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4

    def _info_row(
        self, ws, row: int, label: str, value: Any, *, end_column: int = 12
    ) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=end_column)
        ws.cell(row, 1, label)
        ws.cell(row, 3, value)
        ws.cell(row, 1).font = Font(bold=True, color=self.colors["navy"])
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=self.colors["pale_blue"])
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=self.colors["gray"])
        for cell in ws[row]:
            cell.border = self._border()
        ws.cell(row, 3).alignment = Alignment(wrap_text=True)

    def _title_style(self, cell) -> None:
        cell.font = Font(size=22, bold=True, color=self.colors["white"])
        cell.fill = PatternFill("solid", fgColor=self.colors["navy"])
        cell.alignment = Alignment(vertical="center", horizontal="left")

    def _section_style(self, cell) -> None:
        cell.font = Font(size=13, bold=True, color=self.colors["white"])
        cell.fill = PatternFill("solid", fgColor=self.colors["blue"])
        cell.alignment = Alignment(vertical="center")

    def _image_section_style(self, cell, database: str) -> None:
        fills = {
            "Franklin": "D9EAF7",
            "OncoKB": "EAF5ED",
            "MTBP": "FFF3E8",
        }
        cell.font = Font(size=13, bold=True, color=self.colors["navy"])
        cell.fill = PatternFill("solid", fgColor=fills.get(database, self.colors["gray"]))

    def _header_style(self, cell) -> None:
        cell.fill = PatternFill("solid", fgColor=self.colors["navy"])
        cell.font = Font(bold=True, color=self.colors["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = self._border()

    def _write_headers(self, ws, headers: list[str]) -> None:
        for column, header in enumerate(headers, start=1):
            cell = ws.cell(1, column, header)
            self._header_style(cell)

    def _fit_columns(self, ws, max_width: int) -> None:
        for column_cells in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max(width, 11), max_width
            )

    def _border(self) -> Border:
        side = Side(style="thin", color=self.colors["border"])
        return Border(left=side, right=side, top=side, bottom=side)

    def _variant_sheet_name(
        self,
        variant: VariantRecord,
        index: int,
        used_names: set[str],
        *,
        duplicate_gene: bool,
    ) -> str:
        gene = variant.symbol or f"Variant {index}"
        detail = ""
        if duplicate_gene:
            detail = (variant.hgvsp or "").split(":", 1)[-1]
            if not detail:
                detail = (variant.hgvsc or "").split(":", 1)[-1]
        base = f"{gene} {detail}".strip()
        name = re.sub(r"[\\/*?:\[\]]+", "_", base)[:31]
        suffix = 2
        used_casefold = {used.casefold() for used in used_names}
        while name.casefold() in used_casefold:
            suffix_text = f"-{suffix}"
            name = f"{base[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        return name

    def _compact_evidence(self, items: list[DatabaseEvidence]) -> str:
        if not items:
            return "Ikke søkt"
        item = next((candidate for candidate in items if candidate.status == "found"), items[0])
        if item.clinical_significance:
            label = item.clinical_significance.strip()
        else:
            for pattern in (
                r"classification=([^;|]+)",
                r"oncogenic(?:ity)?=([^;|]+)",
                r"evidence=(Evidence\s+[A-Z](?:\s*\([^)]+\))?(?:,\s*Evidence\s+[A-Z](?:\s*\([^)]+\))?)*)",
                r"functional_relevance=([^;|]+)",
            ):
                match = re.search(pattern, item.summary, flags=re.IGNORECASE)
                if match:
                    label = match.group(1).strip()
                    break
            else:
                status_labels = {
                    "found": "Funnet",
                    "not_found": "Ikke funnet",
                    "invalid_query": "Mangler søkegrunnlag",
                    "login_required": "Innlogging kreves",
                    "rate_limited": "Ratebegrenset",
                    "error": "Feil",
                    "ambiguous_result": "Tvetydig resultat",
                }
                label = status_labels.get(item.status, item.status.replace("_", " ").title())
        return self._append_mtbp_functional_evidence(item, label)

    @staticmethod
    def _mtbp_functional_evidence_label(item: DatabaseEvidence) -> str:
        """Normalized 'Evidence A (Curated)'-style label from an MTBP capture."""
        sources = [item.summary or "", str(item.raw.get("functional_evidence") or "")]
        for source in sources:
            match = re.search(
                r"Evidence\s+[A-Z](?:\s*\([^)]*\))?",
                source,
                flags=re.IGNORECASE,
            )
            if match:
                return " ".join(match.group(0).split())
        return ""

    def _append_mtbp_functional_evidence(
        self, item: DatabaseEvidence, label: str
    ) -> str:
        """Surface 'Evidence A (Curated)' next to the MTBP classification."""
        if item.database != "MTBP":
            return label
        evidence_label = self._mtbp_functional_evidence_label(item)
        if not evidence_label or evidence_label.casefold() in label.casefold():
            return label
        return f"{label} – {evidence_label}"

    def _text_evidence(self, items: list[DatabaseEvidence]) -> str:
        if not items:
            return "Not queried or no evidence record available."
        return "\n\n".join(
            " | ".join(
                part
                for part in [
                    f"Status: {item.status.replace('_', ' ').title()}",
                    f"Classification: {item.clinical_significance}"
                    if item.clinical_significance
                    else "",
                    f"Accession: {item.accession}" if item.accession else "",
                    item.summary,
                    f"Source: {item.url}"
                    if item.url and item.database != "MTBP"
                    else "",
                ]
                if part
            )
            for item in items
        )

    def _status_text(self, items: list[DatabaseEvidence]) -> str:
        if not items:
            return "Not queried"
        return ", ".join(
            dict.fromkeys(item.status.replace("_", " ").title() for item in items)
        )

    def _screenshot_records(
        self, items: list[DatabaseEvidence]
    ) -> list[dict[str, str]]:
        records: list[dict[str, str]] = []
        seen: set[str] = set()
        for item in items:
            raw_records = item.raw.get("screenshots")
            if isinstance(raw_records, list):
                for record in raw_records:
                    if not isinstance(record, dict):
                        continue
                    path = str(record.get("path") or "").strip()
                    if not path or path in seen:
                        continue
                    seen.add(path)
                    records.append(
                        {
                            "label": str(record.get("label") or f"{item.database} evidence"),
                            "path": path,
                            "url": ""
                            if item.database == "MTBP"
                            else str(record.get("url") or item.url or ""),
                        }
                    )
            legacy = str(item.raw.get("screenshot") or "").strip()
            if legacy and legacy not in seen:
                seen.add(legacy)
                records.append(
                    {
                        "label": f"{item.database} evidence",
                        "path": legacy,
                        "url": "" if item.database == "MTBP" else item.url,
                    }
                )
        return records

    def _by_database(
        self, items: list[DatabaseEvidence]
    ) -> dict[str, list[DatabaseEvidence]]:
        grouped: dict[str, list[DatabaseEvidence]] = defaultdict(list)
        for item in items:
            if item.database in REPORT_DATABASES:
                grouped[item.database].append(item)
        return grouped

    @staticmethod
    def _cell_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    @staticmethod
    def _key(variant: VariantRecord) -> str:
        return f"{variant.sample}|{variant.hgvsc}"
