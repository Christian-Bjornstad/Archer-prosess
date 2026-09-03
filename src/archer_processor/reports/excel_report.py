from __future__ import annotations

import os
from collections import defaultdict
from math import isnan
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from archer_processor.core.highlights import priority_warning, variant_highlight
from archer_processor.core.models import DatabaseEvidence, ProcessingResult, VariantRecord
from archer_processor.core.sorting import variant_sort_key


DEFAULT_DATABASE_COLUMNS = [
    "ClinVar",
    "MTBP",
    "Franklin",
    "OncoKB",
    "COSMIC",
]

REFERENCE_HIDDEN_COLUMNS = {
    "Report",
    "Primer Alt Reads +",
    "Primer Alt Reads -",
    "Primer Ref Reads +",
    "Primer Ref Reads -",
    "UDP",
    "DRO",
    "URO",
    "Seq Alt Reads +",
    "Seq Alt Reads -",
    "Seq Ref Reads +",
    "Seq Ref Reads -",
    "Variant Name",
    "Variant Call",
    "FATHMM",
    "RO",
    "Has Seq Dir Bias",
    "MA Match",
    "Quality Rating",
    "FEDSP",
    "FEDSR",
    "Intra Job BN",
    "Intra Job MDAF",
    "Intra Job AOC",
    "Intra Job AFC",
    "ND BN",
    "ND MDAF",
    "ND AOC",
    "ND AFC",
    "ND DAF Outlier P Value",
    "ND DBN",
    "ND DMDAF",
    "ND D95MDAF",
    "ND DAOC",
    "ND DAFC",
    "ND SAF Outlier P Value",
    "ND SBN",
    "ND SMDAF",
    "ND S95MDAF",
    "ND SAOC",
    "ND SAFC",
    "Min Outlier P Value",
    "Intra Job DBN",
    "Intra Job DAF Outlier P Value",
    "Intra Job DMDAF",
    "Intra Job D95MDAF",
    "Intra Job DAOC",
    "Intra Job DAFC",
    "Intra Job SAF Outlier P Value",
    "Intra Job SBN",
    "Intra Job SMDAF",
    "Intra Job S95MDAF",
    "Intra Job SAOC",
    "Intra Job SAFC",
    "AF Outlier P Value",
    "95MDAF",
}


class ExcelReportWriter:
    colors = {
        "navy": "163B5C",
        "blue": "2F75B5",
        "pale_blue": "D9EAF7",
        "green": "E2F0D9",
        "strong_green": "C6EFCE",
        "weak_green": "E9F6EF",
        "yellow": "FFFF00",
        "orange": "FFC000",
        "light_orange": "F4B183",
        "red": "C00000",
        "red_text": "FFFFFF",
        "gray": "F2F2F2",
        "border": "B7C9D6",
        "white": "FFFFFF",
    }

    def write(
        self,
        result: ProcessingResult,
        output_path: Path,
        evidence: dict[str, list[DatabaseEvidence]] | None = None,
        hide_excluded: bool = False,
        database_skip_keys: set[str] | None = None,
    ) -> Path:
        evidence = evidence or {}
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._raw_variant_sheet(
            workbook,
            "With Artifacts",
            result.variants,
            evidence,
            hide_excluded,
            include_selection=True,
            database_skip_keys=database_skip_keys or set(),
        )
        self._raw_variant_sheet(
            workbook,
            "Artifacts Removed",
            [
                variant
                for variant in result.variants
                if variant_highlight(variant) not in {"artifact", "artifact_light"}
            ],
            evidence,
            hide_excluded,
        )
        for ws in workbook.worksheets:
            ws.sheet_view.showGridLines = False
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _database_selection_sheet(
        self,
        workbook: Workbook,
        variants: list[VariantRecord],
        skip_keys: set[str],
    ) -> None:
        """Editable review queue: mark X only for variants that should be skipped."""
        ws = workbook.create_sheet("Database Selection")
        ws.sheet_properties.tabColor = self.colors["yellow"]
        headers = [
            "Skip Database Search (X)",
            "Patient",
            "Sample",
            "Gene",
            "HGVSc",
            "HGVSp",
            "AF",
            "Depth",
            "Decision",
            "Decision Reason",
            "Consequence",
            "Clinical Significance",
            "History",
            "Warnings",
        ]
        self._headers(ws, headers)
        ws["A1"].fill = PatternFill("solid", fgColor=self.colors["yellow"])
        ws["A1"].font = Font(bold=True, color=self.colors["navy"])
        sorted_variants = sorted(variants, key=variant_sort_key)
        for row_index, variant in enumerate(sorted_variants, start=2):
            values = [
                "X" if self._key(variant) in skip_keys else "",
                variant.patient_id,
                variant.sample,
                variant.symbol,
                variant.hgvsc,
                variant.hgvsp,
                variant.af,
                variant.depth,
                variant.decision,
                variant.decision_reason,
                variant.consequence,
                variant.clinical_significance,
                "\n".join(str(item) for item in variant.history_matches),
                "\n".join(
                    value
                    for value in [*variant.warnings, priority_warning(variant)]
                    if value
                ),
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row_index, column, value)
                cell.border = self._border()
                cell.alignment = Alignment(
                    vertical="top", wrap_text=column in {10, 11, 12, 13, 14}
                )
            ws.cell(row_index, 1).fill = PatternFill(
                "solid", fgColor=self.colors["yellow"]
            )
            ws.cell(row_index, 1).alignment = Alignment(horizontal="center")
            if variant.af is not None:
                ws.cell(row_index, 7).number_format = "0.00%"
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:N{max(1, len(variants) + 1)}"
        self._fit(ws, max_width=48)
        ws.column_dimensions["A"].width = 25

    def _summary_sheet(
        self,
        workbook: Workbook,
        result: ProcessingResult,
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> None:
        ws = workbook.create_sheet("Summary")
        ws.sheet_properties.tabColor = self.colors["navy"]
        ws.merge_cells("A1:H2")
        ws["A1"] = "Archer VPM Variant Review"
        ws["A1"].font = Font(size=22, bold=True, color=self.colors["white"])
        ws["A1"].fill = PatternFill("solid", fgColor=self.colors["navy"])
        ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
        ws.merge_cells("A3:H3")
        ws["A3"] = (
            "Decision overview and database evidence • verify source pages before "
            "clinical use • MTBP public output is research-only"
        )
        ws["A3"].font = Font(italic=True, color="5E6A73")
        ws["A3"].alignment = Alignment(wrap_text=True)

        evidence_items = [item for items in evidence.values() for item in items]
        cards = [
            ("A5:B5", "A6:B7", "Total variants", result.total_count, self.colors["pale_blue"]),
            ("C5:D5", "C6:D7", "Included", len(result.included), self.colors["green"]),
            ("E5:F5", "E6:F7", "Excluded", len(result.excluded), self.colors["orange"]),
            ("G5:H5", "G6:H7", "Evidence records", len(evidence_items), self.colors["yellow"]),
        ]
        for label_range, value_range, label, value, fill in cards:
            ws.merge_cells(label_range)
            ws.merge_cells(value_range)
            label_cell = ws[label_range.split(":", 1)[0]]
            value_cell = ws[value_range.split(":", 1)[0]]
            label_cell.value = label
            value_cell.value = value
            for cell in (label_cell, value_cell):
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            label_cell.font = Font(bold=True, color=self.colors["navy"])
            value_cell.font = Font(size=20, bold=True, color=self.colors["navy"])

        ws["A10"] = "Run information"
        ws["A10"].font = Font(size=13, bold=True, color=self.colors["navy"])
        rows = [
            ("Input file", str(result.input_path)),
            ("Run date", result.run_date),
            ("Processing time", f"{result.duration_seconds:.2f} s"),
            (
                "Sources queried",
                ", ".join(dict.fromkeys(item.database for item in evidence_items))
                if evidence_items else "None yet",
            ),
        ]
        for row_index, (label, value) in enumerate(rows, start=11):
            ws.cell(row_index, 1, label).font = Font(bold=True, color=self.colors["navy"])
            ws.cell(row_index, 2, value)
            ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
            ws.cell(row_index, 2).alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )

        ws["A18"] = "Evidence status legend"
        ws["A18"].font = Font(size=13, bold=True, color=self.colors["navy"])
        legend = [
            ("Found", "Provider returned a matching result", self.colors["green"]),
            ("Not found", "No matching evidence was returned", self.colors["gray"]),
            ("Review needed", "Login, timeout, ambiguity, or provider error", self.colors["orange"]),
        ]
        for row_index, (label, meaning, fill) in enumerate(legend, start=19):
            ws.cell(row_index, 1, label)
            ws.cell(row_index, 1).font = Font(bold=True)
            ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=fill)
            ws.cell(row_index, 2, meaning)
            ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=5)
        ws.column_dimensions["A"].width = 22
        for column in "BCDEFGH":
            ws.column_dimensions[column].width = 15
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = "A10"

    def _included_sheet(
        self,
        workbook: Workbook,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
    ) -> None:
        ws = workbook.create_sheet("Included Variants")
        ws.sheet_properties.tabColor = "4F8A5B"
        headers = [
            "Sample", "Gene", "HGVSc", "HGVSp", "Transcript", "AF", "Depth",
            "Decision", "Classification", "Evidence Sources", "Evidence Summary",
        ]
        self._headers(ws, headers)
        for row_index, variant in enumerate(variants, start=2):
            items = evidence.get(self._key(variant), [])
            sources = ", ".join(dict.fromkeys(item.database for item in items))
            findings = "\n".join(
                f"{item.database} [{item.status.replace('_', ' ')}]: "
                f"{item.clinical_significance or item.summary}".strip()
                for item in items
            )
            values = [
                variant.sample, variant.symbol, variant.hgvsc, variant.hgvsp,
                variant.transcript, variant.af, variant.depth, variant.decision,
                variant.classification, sources, findings,
            ]
            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row_index, col_index, value)
                cell.border = self._border()
                cell.alignment = Alignment(vertical="top", wrap_text=col_index in {9, 10, 11})
                if col_index == 6 and value is not None:
                    cell.number_format = "0.00%"
            if findings:
                ws.row_dimensions[row_index].height = min(90, 20 + findings.count("\n") * 16)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:K{max(1, len(variants) + 1)}"
        self._fit(ws, max_width=55)

    def _variant_sheet(
        self,
        workbook: Workbook,
        title: str,
        variants: list[VariantRecord],
        run_date: str,
        evidence: dict[str, list[DatabaseEvidence]],
        hide_excluded: bool,
    ) -> None:
        ws = workbook.create_sheet(title)
        database_columns = self._database_columns(evidence)
        headers = [
            "Sample",
            "Patient",
            "Decision",
            "Reason",
            "Gene",
            "HGVSc",
            "HGVSp",
            "Transcript",
            "AF",
            "Depth",
            "AO",
            "Type",
            "Quality",
            "Consequence",
            "Genomic Location",
            "Ref",
            "Alt",
            "Classification",
            "Report",
            "Artifact",
            "History",
            "Warnings",
            *[f"{database} Evidence" for database in database_columns],
            "Run Date",
        ]
        self._headers(ws, headers)
        for row_index, variant in enumerate(variants, start=2):
            evidence_by_database = self._evidence_by_database(evidence.get(self._key(variant), []))
            values = [
                variant.sample,
                variant.patient_id,
                variant.decision,
                variant.decision_reason,
                variant.symbol,
                variant.hgvsc,
                variant.hgvsp,
                variant.transcript,
                variant.af,
                variant.depth,
                variant.ao,
                variant.variant_type,
                variant.quality_score,
                variant.consequence,
                variant.genomic_location,
                variant.ref_allele,
                variant.alt_allele,
                variant.classification,
                variant.report_status,
                variant.artifact_status,
                f"{len(variant.history_matches)} previous match(es)",
                "; ".join(
                    value
                    for value in [*variant.warnings, priority_warning(variant)]
                    if value
                ),
                *[self._evidence_cell(evidence_by_database.get(database, [])) for database in database_columns],
                run_date,
            ]
            evidence_start = headers.index(f"{database_columns[0]} Evidence") + 1 if database_columns else 0
            evidence_columns = set(range(evidence_start, evidence_start + len(database_columns)))
            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row_index, col_index, value)
                cell.border = self._border()
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=col_index in {4, 14, 21, 22} or col_index in evidence_columns,
                )
                if col_index == 9 and value is not None:
                    cell.number_format = "0.00%"
            self._style_variant_row(ws, row_index, variant)
            if hide_excluded and variant.decision == "excluded":
                ws.row_dimensions[row_index].hidden = True
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(variants) + 1}"
        self._fit(ws, max_width=46)

    def _history_sheet(self, workbook: Workbook, variants: list[VariantRecord]) -> None:
        ws = workbook.create_sheet("Local History")
        headers = ["Sample", "Gene", "HGVSc", "Previous Sample", "Previous Classification", "TO", "Rept", "Germ", "Artf"]
        self._headers(ws, headers)
        row_index = 2
        for variant in variants:
            for match in variant.history_matches:
                values = [
                    variant.sample,
                    variant.symbol,
                    variant.hgvsc,
                    match.get("Sample"),
                    match.get("Classification"),
                    match.get("TO"),
                    match.get("Rept"),
                    match.get("Germ"),
                    match.get("Artf"),
                ]
                for col_index, value in enumerate(values, start=1):
                    ws.cell(row_index, col_index, value).border = self._border()
                row_index += 1
        self._fit(ws)

    def _database_sheet(
        self,
        workbook: Workbook,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
        report_directory: Path,
    ) -> None:
        ws = workbook.create_sheet("Database Evidence")
        ws.sheet_properties.tabColor = self.colors["blue"]
        headers = [
            "Sample", "Gene", "HGVSc", "Database", "Status", "Significance",
            "Accession", "Summary", "Source Page", "Screenshot", "Captured At",
        ]
        self._headers(ws, headers)
        row_index = 2
        for variant in variants:
            for item in evidence.get(self._key(variant), []):
                if item.database not in DEFAULT_DATABASE_COLUMNS:
                    continue
                values = [
                    variant.sample,
                    variant.symbol,
                    variant.hgvsc,
                    item.database,
                    item.status.replace("_", " ").title(),
                    item.clinical_significance,
                    item.accession,
                    item.summary,
                    "Open source" if item.url else "",
                    "Open screenshot" if item.raw.get("screenshot") else "",
                    item.raw.get("captured_at", ""),
                ]
                for col_index, value in enumerate(values, start=1):
                    cell = ws.cell(row_index, col_index, value)
                    cell.border = self._border()
                    cell.alignment = Alignment(vertical="top", wrap_text=col_index in {6, 8})
                if item.url:
                    ws.cell(row_index, 9).hyperlink = item.url
                    ws.cell(row_index, 9).style = "Hyperlink"
                screenshot = str(item.raw.get("screenshot") or "")
                if screenshot and Path(screenshot).exists():
                    try:
                        screenshot_target = os.path.relpath(
                            Path(screenshot).resolve(), report_directory.resolve()
                        )
                    except ValueError:
                        screenshot_target = Path(screenshot).resolve().as_uri()
                    ws.cell(row_index, 10).hyperlink = screenshot_target
                    ws.cell(row_index, 10).style = "Hyperlink"
                status_fill = {
                    "found": self.colors["green"],
                    "not_found": self.colors["gray"],
                }.get(item.status, self.colors["orange"])
                ws.cell(row_index, 5).fill = PatternFill("solid", fgColor=status_fill)
                ws.cell(row_index, 5).font = Font(bold=True, color=self.colors["navy"])
                ws.row_dimensions[row_index].height = 32
                row_index += 1
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:K{max(1, row_index - 1)}"
        self._fit(ws, max_width=60)

    def _rules_sheet(self, workbook: Workbook, result: ProcessingResult) -> None:
        ws = workbook.create_sheet("Rules")
        self._headers(ws, ["Rule ID"])
        for row_index, rule in enumerate(result.rules_applied, start=2):
            ws.cell(row_index, 1, rule)
        self._fit(ws)

    def _style_variant_row(self, ws, row_index: int, variant: VariantRecord) -> None:
        fill = {
            "artifact": self.colors["orange"],
            "artifact_light": self.colors["light_orange"],
            "germline": self.colors["strong_green"],
            "germline_low_af": self.colors["weak_green"],
        }.get(variant_highlight(variant))
        if fill:
            for cell in ws[row_index]:
                cell.fill = PatternFill("solid", fgColor=fill)

    def _headers(self, ws, headers: list[str]) -> None:
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_index, header)
            cell.fill = PatternFill("solid", fgColor=self.colors["navy"])
            cell.font = Font(bold=True, color=self.colors["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._border()

    def _fit(self, ws, max_width: int = 38) -> None:
        for column_cells in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), max_width)

    def _border(self) -> Border:
        side = Side(style="thin", color=self.colors["border"])
        return Border(left=side, right=side, top=side, bottom=side)

    def _key(self, variant: VariantRecord) -> str:
        return f"{variant.sample}|{variant.hgvsc}"

    def _raw_variant_sheet(
        self,
        workbook: Workbook,
        title: str,
        variants: list[VariantRecord],
        evidence: dict[str, list[DatabaseEvidence]],
        hide_excluded: bool = False,
        *,
        include_selection: bool = False,
        database_skip_keys: set[str] | None = None,
    ) -> None:
        ws = workbook.create_sheet(title)
        raw_columns = self._raw_columns(variants)
        database_columns = self._database_columns(evidence)
        headers = (
            (["Skip Database Search (X)"] if include_selection else [])
            + raw_columns
            + [f"{database} Evidence" for database in database_columns]
        )
        self._headers(ws, headers)
        raw_offset = 1 if include_selection else 0
        evidence_start = raw_offset + len(raw_columns) + 1
        evidence_columns = set(range(evidence_start, evidence_start + len(database_columns)))
        skip_keys = database_skip_keys or set()

        sorted_variants = sorted(variants, key=variant_sort_key)
        for row_index, variant in enumerate(sorted_variants, start=2):
            evidence_by_database = self._evidence_by_database(evidence.get(self._key(variant), []))
            values = [
                *(
                    [
                        "X"
                        if self._key(variant) in skip_keys
                        or variant_highlight(variant) in {"artifact", "artifact_light"}
                        else ""
                    ]
                    if include_selection
                    else []
                ),
                *[
                    self._raw_value(
                        variant.af if column == "AF" else variant.raw.get(column)
                    )
                    for column in raw_columns
                ],
                *[self._evidence_cell(evidence_by_database.get(database, [])) for database in database_columns],
            ]
            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row_index, col_index, value)
                cell.border = self._border()
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                raw_index = col_index - raw_offset - 1
                if 0 <= raw_index < len(raw_columns) and raw_columns[raw_index] == "AF" and value not in [None, ""]:
                    cell.number_format = "0.00%"
            if include_selection:
                ws.cell(row_index, 1).fill = PatternFill(
                    "solid", fgColor=self.colors["yellow"]
                )
                ws.cell(row_index, 1).font = Font(bold=True, color=self.colors["navy"])
                ws.cell(row_index, 1).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            self._style_variant_row(ws, row_index, variant)
            if include_selection:
                ws.cell(row_index, 1).fill = PatternFill(
                    "solid", fgColor=self.colors["yellow"]
                )
            ws.row_dimensions[row_index].height = 18
            if hide_excluded and variant.decision == "excluded":
                ws.row_dimensions[row_index].hidden = True
        ws.freeze_panes = "G2" if include_selection else "F2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(variants) + 1}"
        self._fit(ws, max_width=46)
        if include_selection:
            ws.column_dimensions["A"].width = 25
            ws["A1"].fill = PatternFill("solid", fgColor=self.colors["yellow"])
            ws["A1"].font = Font(bold=True, color=self.colors["navy"])
        for index, header in enumerate(headers, start=1):
            letter = get_column_letter(index)
            if header in REFERENCE_HIDDEN_COLUMNS:
                ws.column_dimensions[letter].hidden = True
            elif header.endswith(" Evidence"):
                ws.column_dimensions[letter].width = 34
        ws.row_dimensions[1].height = 34

    def _raw_columns(self, variants: list[VariantRecord]) -> list[str]:
        columns: list[str] = []
        for variant in variants:
            for column in variant.raw:
                if column not in columns:
                    columns.append(column)
        return columns

    def _raw_value(self, value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, float) and isnan(value):
            return ""
        return value

    def _database_columns(self, evidence: dict[str, list[DatabaseEvidence]]) -> list[str]:
        return list(DEFAULT_DATABASE_COLUMNS)

    def _evidence_by_database(self, evidence_items: list[DatabaseEvidence]) -> dict[str, list[DatabaseEvidence]]:
        grouped: dict[str, list[DatabaseEvidence]] = defaultdict(list)
        for item in evidence_items:
            grouped[item.database].append(item)
        return grouped

    def _evidence_cell(self, evidence_items: list[DatabaseEvidence]) -> str:
        return "\n".join(
            f"[{item.status}] {item.summary}".strip()
            for item in evidence_items
        )
