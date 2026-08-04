from __future__ import annotations

from pathlib import Path

import openpyxl


SELECTION_SHEET = "Database Selection"
SKIP_HEADER = "Skip Database Search (X)"
SAMPLE_HEADER = "Sample"
HGVSC_HEADER = "HGVSc"


def load_database_skip_keys(workbook_path: Path) -> set[str]:
    """Return variant keys explicitly marked X in a processed review workbook."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if SELECTION_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Workbook does not contain the '{SELECTION_SHEET}' sheet. "
                "Create it with the current Archer Prosess version first."
            )
        worksheet = workbook[SELECTION_SHEET]
        headers = {
            str(cell.value or "").strip(): index
            for index, cell in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1)))
        }
        if any(
            header not in headers
            for header in (SKIP_HEADER, SAMPLE_HEADER, HGVSC_HEADER)
        ):
            raise ValueError(
                f"'{SELECTION_SHEET}' is missing its Sample, HGVSc, or skip column."
            )
        skip_index = headers[SKIP_HEADER]
        sample_index = headers[SAMPLE_HEADER]
        hgvsc_index = headers[HGVSC_HEADER]
        skipped: set[str] = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            sample = str(row[sample_index] or "").strip()
            hgvsc = str(row[hgvsc_index] or "").strip()
            marker = str(row[skip_index] or "").strip().casefold()
            if sample and hgvsc and marker == "x":
                skipped.add(f"{sample}|{hgvsc}")
        return skipped
    finally:
        workbook.close()
