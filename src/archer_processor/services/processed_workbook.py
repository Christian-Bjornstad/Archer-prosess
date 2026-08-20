from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from collections.abc import Callable
from typing import Any

import openpyxl

from archer_processor.core.models import DatabaseEvidence, ProcessingResult, VariantRecord
from archer_processor.core.rules import FilterEngine
from archer_processor.io import ArcherTsvReader

from .database_selection import HGVSC_HEADER, SAMPLE_HEADER, SELECTION_SHEET, SKIP_HEADER
from .evidence_audit import EvidenceAuditIndex, migrate_loaded_evidence


@dataclass(slots=True)
class ProcessedWorkbookState:
    result: ProcessingResult
    evidence: dict[str, list[DatabaseEvidence]]
    database_skip_keys: set[str]


class ProcessedWorkbookLoader:
    """Restore a VPM review session from its two-sheet processed workbook."""

    def __init__(
        self,
        *,
        filter_engine: FilterEngine | None = None,
    ) -> None:
        self.reader = ArcherTsvReader()
        self.filter_engine = filter_engine or FilterEngine()

    def load(
        self,
        workbook_path: Path,
        progress: Callable[[int, int, str], None] | None = None,
    ) -> ProcessedWorkbookState:
        workbook_path = workbook_path.resolve()
        if not workbook_path.exists():
            raise ValueError(f"Workbook not found: {workbook_path}")
        if workbook_path.suffix.casefold() != ".xlsx":
            raise ValueError("Select a processed VPM workbook in .xlsx format.")

        workbook = openpyxl.load_workbook(
            workbook_path, read_only=True, data_only=True
        )
        try:
            if SELECTION_SHEET not in workbook.sheetnames:
                raise ValueError(
                    f"This is not a current VPM review workbook: the "
                    f"'{SELECTION_SHEET}' sheet is missing."
                )
            worksheet = workbook[SELECTION_SHEET]
            header_values = [
                str(cell.value or "").strip()
                for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
            ]
            if len(header_values) != len(set(header_values)):
                raise ValueError("The processed workbook contains duplicate column names.")
            headers = {header: index for index, header in enumerate(header_values)}
            missing = sorted(
                (self.reader.required_columns | {SKIP_HEADER}) - set(headers)
            )
            if missing:
                raise ValueError(
                    "The processed workbook is missing required columns: "
                    + ", ".join(missing)
                )

            evidence_headers = {
                header.removesuffix(" Evidence"): index
                for header, index in headers.items()
                if header.endswith(" Evidence")
            }
            raw_headers = [
                header
                for header in header_values
                if header != SKIP_HEADER and not header.endswith(" Evidence")
            ]
            variants: list[VariantRecord] = []
            total_rows = max(1, worksheet.max_row - 1)
            skip_keys: set[str] = set()
            cell_evidence: dict[str, dict[str, str]] = {}
            for source_row, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                sample = self._text(row[headers[SAMPLE_HEADER]])
                hgvsc = self._text(row[headers[HGVSC_HEADER]])
                if not sample and not hgvsc:
                    continue
                raw = {
                    header: row[headers[header]]
                    for header in raw_headers
                }
                variant = self.reader.row_to_variant(
                    workbook_path, source_row, raw
                )
                if not variant.sample:
                    raise ValueError(
                        f"Row {source_row} is missing Sample and cannot be restored."
                    )
                variants.append(variant)
                if progress and (len(variants) == 1 or len(variants) % 25 == 0):
                    progress(
                        len(variants), total_rows, f"Reading variant {len(variants)}"
                    )
                key = self._variant_key(variant)
                if self._text(row[headers[SKIP_HEADER]]).casefold() == "x":
                    skip_keys.add(key)
                cell_evidence[key] = {
                    database: self._text(row[column_index])
                    for database, column_index in evidence_headers.items()
                }
        finally:
            workbook.close()

        if not variants:
            raise ValueError("The processed workbook does not contain any variants.")
        self.filter_engine.apply(variants)
        evidence = self._restore_evidence(
            workbook_path, variants, cell_evidence, progress=progress
        )
        timestamp = datetime.fromtimestamp(workbook_path.stat().st_mtime)
        result = ProcessingResult(
            input_path=workbook_path,
            output_path=workbook_path,
            run_date=timestamp.date().isoformat(),
            variants=variants,
            rules_applied=[rule.rule_id for rule in self.filter_engine.rules],
            started_at=timestamp,
            finished_at=timestamp,
        )
        return ProcessedWorkbookState(result, evidence, skip_keys)

    def _restore_evidence(
        self,
        workbook_path: Path,
        variants: list[VariantRecord],
        cell_evidence: dict[str, dict[str, str]],
        *,
        progress: Callable[[int, int, str], None] | None = None,
    ) -> dict[str, list[DatabaseEvidence]]:
        artifact_root = workbook_path.parent / f"{workbook_path.stem}_browser_evidence"
        audit_index = EvidenceAuditIndex.build(artifact_root)
        patient_order = list(dict.fromkeys(variant.patient_id for variant in variants))
        patient_indexes = {
            patient_id: index
            for index, patient_id in enumerate(patient_order, start=1)
        }
        restored: dict[str, list[DatabaseEvidence]] = {}
        for index, variant in enumerate(variants, start=1):
            key = self._variant_key(variant)
            databases = cell_evidence.get(key, {})
            items: list[DatabaseEvidence] = []
            for database, cell_value in databases.items():
                audit = self._load_audit(
                    audit_index,
                    artifact_root,
                    patient_indexes[variant.patient_id],
                    database,
                    variant,
                )
                if audit is not None:
                    items.append(audit)
                elif cell_value:
                    items.extend(self._parse_evidence_cell(database, cell_value))
            if items:
                restored[key] = items
            if progress:
                progress(index, len(variants), f"Restoring evidence {index}/{len(variants)}")
        return restored

    def _load_audit(
        self,
        audit_index: EvidenceAuditIndex,
        artifact_root: Path,
        patient_index: int,
        database: str,
        variant: VariantRecord,
    ) -> DatabaseEvidence | None:
        selected = audit_index.best(database, variant)
        if selected is None:
            return None
        _, payload = selected
        raw = payload.get("raw")
        evidence = DatabaseEvidence(
            database=database,
            status=self._text(payload.get("status")) or "found",
            summary=self._text(payload.get("summary")),
            accession=self._text(payload.get("accession")),
            clinical_significance=self._text(payload.get("clinical_significance")),
            url=self._text(payload.get("url")),
            raw=raw if isinstance(raw, dict) else {},
        )
        return migrate_loaded_evidence(evidence, artifact_root)

    @staticmethod
    def _parse_evidence_cell(
        database: str, value: str
    ) -> list[DatabaseEvidence]:
        starts = list(re.finditer(r"(?m)^\[([^\]]+)]\s*", value))
        if not starts:
            return [DatabaseEvidence(database, "restored", value)]
        items: list[DatabaseEvidence] = []
        for index, match in enumerate(starts):
            end = starts[index + 1].start() if index + 1 < len(starts) else len(value)
            summary = value[match.end():end].strip()
            items.append(
                DatabaseEvidence(database, match.group(1).strip(), summary)
            )
        return items

    @staticmethod
    def _variant_key(variant: VariantRecord) -> str:
        return f"{variant.sample}|{variant.hgvsc}"

    @staticmethod
    def _text(value: Any) -> str:
        return "" if value is None else str(value).strip()
